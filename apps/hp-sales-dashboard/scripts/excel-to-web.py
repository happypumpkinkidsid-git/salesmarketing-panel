#!/usr/bin/env python3
"""
excel-to-web.py — Sync KOL Pool data from Excel → js/kol-pool-data.js

Usage:
    python3 scripts/excel-to-web.py
    python3 scripts/excel-to-web.py --excel /path/to/DEPLOY-KOL-Dashboard.xlsx
    python3 scripts/excel-to-web.py --zip   (also creates hp-sales-dashboard-latest.zip)

Reads:  Projects/DEPLOY-KOL-Dashboard.xlsx  →  "🌊 KOL Pool" sheet
Writes: js/kol-pool-data.js

Column mapping (Excel KOL Pool → JS schema):
  A  Handle / IG          → handle
  B  Nama Lengkap         → nama
  C  Platform             → platform
  D  Tier                 → tier
  E  Niche                → niche
  F  Followers            → followers
  G  Avg Views            → avg_views
  H  ER %                 → er_persen
  I  Rate Diminta         → rate_diminta
  J  Rate Deal            → rate_deal
  K  CPM (Rp)             → cpm
  L  CPM Zone             → cpm_zone
  M  Status               → status
  N  Scope / Deliverables → scope
  O  Brief Bulan          → brief_bulan
  P  Produk               → produk
  Q  Kontak WA            → kontak_wa
  R  Keputusan            → keputusan
  S  Catatan              → catatan
"""

import sys
import os
import argparse
import subprocess
from datetime import date

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR  = os.path.dirname(SCRIPT_DIR)
PARENT_DIR   = os.path.dirname(PROJECT_DIR)
EXCEL_PATH   = os.path.join(PARENT_DIR, 'DEPLOY-KOL-Dashboard.xlsx')
OUTPUT_PATH  = os.path.join(PROJECT_DIR, 'js', 'kol-pool-data.js')
POOL_SHEET   = '🌊 KOL Pool'
HEADER_ROW   = 1   # row index (0-based) of the header; data starts at row 2

# ── Try to import openpyxl ─────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    venv = '/tmp/kol-venv/bin/pip'
    print('openpyxl not found — installing to /tmp/kol-venv …')
    subprocess.run([sys.executable, '-m', 'venv', '/tmp/kol-venv'], check=True)
    subprocess.run(['/tmp/kol-venv/bin/pip', 'install', 'openpyxl', '-q'], check=True)
    sys.path.insert(0, '/tmp/kol-venv/lib/python%d.%d/site-packages' %
                       (sys.version_info.major, sys.version_info.minor))
    import openpyxl

# ── CPE & CPM helpers ──────────────────────────────────────────────────────────
def cpm_zone(cpm_val):
    if not cpm_val:        return 'unknown'
    if cpm_val <= 5000:    return 'green'
    if cpm_val <= 10000:   return 'blue'
    if cpm_val <= 15000:   return 'amber'
    return 'red'

def safe_int(v):
    try: return int(float(str(v).replace(',', '').replace('Rp', '').strip()))
    except: return 0

def safe_float(v):
    try: return round(float(str(v).replace(',', '.').replace('%', '').strip()), 2)
    except: return 0.0

def safe_str(v):
    if v is None: return ''
    return str(v).strip()

# ── Column indices (0-based, matching the header row) ─────────────────────────
COL = {
    'handle':      0,
    'nama':        1,
    'platform':    2,
    'tier':        3,
    'niche':       4,
    'followers':   5,
    'avg_views':   6,
    'er_persen':   7,
    'rate_diminta':8,
    'rate_deal':   9,
    'cpm':         10,
    'cpm_zone':    11,
    'status':      12,
    'scope':       13,
    'brief_bulan': 14,
    'produk':      15,
    'kontak_wa':   16,
    'keputusan':   17,
    'catatan':     18,
}

VALID_STATUSES = {'KANDIDAT','PENDING','HOLD','NEGOSIASI','DEAL','BARTER','SELESAI','CANCEL'}

def js_str(v):
    return "'" + safe_str(v).replace("'", "\\'") + "'"

def read_pool(wb):
    if POOL_SHEET not in wb.sheetnames:
        raise ValueError(f'Sheet "{POOL_SHEET}" not found. Available: {wb.sheetnames}')
    ws = wb[POOL_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    # skip header row, start from HEADER_ROW+1
    data_rows = rows[HEADER_ROW + 1:]

    records = []
    for i, row in enumerate(data_rows):
        if len(row) == 0:
            continue
        handle = safe_str(row[COL['handle']] if len(row) > COL['handle'] else '')
        if not handle or handle.startswith('#'):
            continue

        def g(key):
            idx = COL[key]
            return row[idx] if len(row) > idx else None

        raw_status = safe_str(g('status')).upper()
        status = raw_status if raw_status in VALID_STATUSES else 'KANDIDAT'

        cpm_val = safe_int(g('cpm'))
        zone    = safe_str(g('cpm_zone')) or cpm_zone(cpm_val)

        rate_d  = safe_int(g('rate_diminta'))
        rate_deal = safe_int(g('rate_deal'))
        avg_v   = safe_int(g('avg_views'))
        cpe_val = round(rate_deal / avg_v, 0) if (rate_deal and avg_v) else 0

        records.append({
            'id':          f'p{i+1}',
            'handle':      handle,
            'nama':        safe_str(g('nama')),
            'platform':    safe_str(g('platform')) or 'Instagram',
            'tier':        safe_str(g('tier')),
            'niche':       safe_str(g('niche')),
            'followers':   safe_int(g('followers')),
            'avg_views':   avg_v,
            'er_persen':   safe_float(g('er_persen')),
            'aqs_score':   0,
            'indonesia_pct': 0,
            'rate_diminta':rate_d,
            'rate_deal':   rate_deal,
            'cpm':         cpm_val,
            'cpm_zone':    zone,
            'cpe':         int(cpe_val),
            'status':      status,
            'scope':       safe_str(g('scope')),
            'brief_bulan': safe_str(g('brief_bulan')),
            'produk':      safe_str(g('produk')),
            'kontak_wa':   safe_str(g('kontak_wa')),
            'sumber':      'Pool',
            'keputusan':   safe_str(g('keputusan')),
            'tgl_tambah':  str(date.today()),
            'tgl_update':  str(date.today()),
            'catatan':     safe_str(g('catatan')),
        })
    return records

def render_js(records):
    today = date.today().isoformat()
    lines = [
        '// ============================================================',
        '// KOL POOL DATA — auto-generated from DEPLOY-KOL-Dashboard.xlsx',
        '// Run: python3 scripts/excel-to-web.py  to regenerate',
        '// DO NOT edit manually — changes will be overwritten on next sync',
        '// ============================================================',
        f'// Last synced: {today}',
        '',
        'const KOL_POOL_DATA = [',
    ]
    for r in records:
        line = (
            f"  {{ id:{js_str(r['id'])}, handle:{js_str(r['handle'])}, "
            f"nama:{js_str(r['nama'])}, platform:{js_str(r['platform'])}, "
            f"tier:{js_str(r['tier'])}, niche:{js_str(r['niche'])}, "
            f"followers:{r['followers']}, avg_views:{r['avg_views']}, "
            f"er_persen:{r['er_persen']}, aqs_score:0, indonesia_pct:0, "
            f"rate_diminta:{r['rate_diminta']}, rate_deal:{r['rate_deal']}, "
            f"cpm:{r['cpm']}, cpm_zone:{js_str(r['cpm_zone'])}, cpe:{r['cpe']}, "
            f"status:{js_str(r['status'])}, scope:{js_str(r['scope'])}, "
            f"brief_bulan:{js_str(r['brief_bulan'])}, produk:{js_str(r['produk'])}, "
            f"kontak_wa:{js_str(r['kontak_wa'])}, sumber:'Pool', "
            f"keputusan:{js_str(r['keputusan'])}, "
            f"tgl_tambah:{js_str(r['tgl_tambah'])}, tgl_update:{js_str(r['tgl_update'])}, "
            f"catatan:{js_str(r['catatan'])} }},"
        )
        lines.append(line)
    lines.append('];')
    lines.append('')
    return '\n'.join(lines)

def main():
    parser = argparse.ArgumentParser(description='Sync Excel KOL Pool → kol-pool-data.js')
    parser.add_argument('--excel', default=EXCEL_PATH, help='Path to DEPLOY-KOL-Dashboard.xlsx')
    parser.add_argument('--zip',   action='store_true', help='Also zip hp-sales-dashboard after sync')
    args = parser.parse_args()

    excel = args.excel
    if not os.path.exists(excel):
        print(f'ERROR: Excel file not found: {excel}')
        print(f'Expected at: {EXCEL_PATH}')
        sys.exit(1)

    print(f'Reading: {excel}')
    wb = openpyxl.load_workbook(excel, data_only=True)
    records = read_pool(wb)
    print(f'Found {len(records)} KOLs in "{POOL_SHEET}"')

    js = render_js(records)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(js)
    print(f'Written: {OUTPUT_PATH}')

    if args.zip:
        import zipfile, shutil
        zip_path = os.path.join(PARENT_DIR, 'hp-sales-dashboard-latest.zip')
        print(f'Zipping → {zip_path}')
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(PROJECT_DIR):
                dirs[:] = [d for d in dirs if d not in {'.git', '.claude', '__pycache__', 'node_modules'}]
                for file in files:
                    if file.endswith('.DS_Store'):
                        continue
                    fp = os.path.join(root, file)
                    arc = os.path.relpath(fp, PARENT_DIR)
                    zf.write(fp, arc)
        print(f'Zip ready: {zip_path}')
        subprocess.run(['open', '-R', zip_path])

    print('Done.')

if __name__ == '__main__':
    main()
